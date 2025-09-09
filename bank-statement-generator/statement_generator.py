import random
from datetime import datetime, timedelta
from pathlib import Path
import sys
import os

# Add the parent directory to the path to import from streamlit_bank_analyzer
sys.path.append(os.path.join(os.path.dirname(__file__), "..", "streamlit_bank_analyzer"))

import excel2img
import pandas as pd
from faker import Faker
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Import comprehensive transaction data
from transaction_data import RECURRING_PAYMENTS, ONE_OFF_PAYMENTS, GERMAN_CITIES

# Configuration
NUM_STATEMENTS = 20
OUTPUT_DIR = Path("bank_statements")
TEMPLATE_FILE = Path("template.xlsx")
GROUND_TRUTH_CSV = OUTPUT_DIR / "ground_truth.csv"

# Template layout constants
START_BALANCE_ROW = 5
TRANSACTION_START_ROW = 6
TRANSACTION_END_ROW = 11
END_BALANCE_ROW = 12
NEXT_BILLING_ROW = 13
MAX_TRANSACTIONS = TRANSACTION_END_ROW - TRANSACTION_START_ROW + 1


def select_recurring_payments(num_transactions):
    """Select recurring payments for the statement."""
    guaranteed = []
    if random.random() < 0.6:  # 60% chance for Vodafone
        vodafone = next(p for p in RECURRING_PAYMENTS if p["partner"] == "Vodafone")
        guaranteed.append(vodafone)

    remaining = [p for p in RECURRING_PAYMENTS if p["partner"] != "Vodafone"]
    num_additional = random.randint(0, min(len(remaining), num_transactions - len(guaranteed)))
    guaranteed.extend(random.sample(remaining, num_additional))

    return guaranteed

def generate_recurring_transaction(payment, fake, current_date):
    """Generate a recurring payment transaction."""
    partner = payment["partner"]
    description = payment["description_template"].format(
        partner=partner,
        contract_id=fake.random_number(digits=8)
    )
    amount = round(payment["base_amount"] + random.uniform(*payment["variation"]), 2)

    return {
        "partner": partner,
        "category": payment["category"],
        "description": description,
        "amount": amount,
        "date": current_date
    }

def generate_one_off_transaction(fake, current_date):
    """Generate a one-off payment transaction."""
    one_off = random.choice(ONE_OFF_PAYMENTS)
    partner = random.choice(one_off["partner_options"])
    description = one_off["description_template"].format(
        date_short=current_date.strftime("%d.%m"),
        partner=partner,
        random_str=fake.lexify(text='??????').upper(),
        city=fake.city()
    )
    amount = round(random.uniform(*one_off["amount_range"]), 2)

    return {
        "partner": partner,
        "category": one_off["category"],
        "description": description,
        "amount": amount,
        "date": current_date
    }

def format_transaction_for_excel(transaction):
    """Format transaction data for Excel output."""
    beleg_date = transaction["date"].strftime("%d.%m.")
    valuta_date = (transaction["date"] + timedelta(days=1)).strftime("%d.%m.")
    amount_str = f"{transaction['amount']:.2f}".replace('.', ',') + "-"

    return [beleg_date, valuta_date, transaction["description"], amount_str]

def generate_statement_data(fake, statement_number):
    """
    Generate realistic transaction data with a mix of recurring and one-off payments.
    """
    start_date = fake.date_between(start_date="-2y", end_date="today")
    current_date = start_date
    num_transactions = random.randint(3, MAX_TRANSACTIONS)

    recurring_payments = select_recurring_payments(num_transactions)
    transaction_types = recurring_payments + [None] * (num_transactions - len(recurring_payments))
    random.shuffle(transaction_types)

    transactions = []
    ground_truth = []
    current_balance = 0.0

    for transaction_type in transaction_types:
        current_date += timedelta(days=random.randint(1, 4))

        if transaction_type:
            transaction = generate_recurring_transaction(transaction_type, fake, current_date)
        else:
            transaction = generate_one_off_transaction(fake, current_date)

        transactions.append(format_transaction_for_excel(transaction))
        current_balance += transaction["amount"]

        ground_truth.append({
            "statement_id": statement_number,
            "date": current_date.strftime("%d.%m.%Y"),
            "description": transaction["description"],
            "amount": -transaction["amount"],
            "identified_partner": transaction["partner"],
            "category": transaction["category"]
        })

    end_date = current_date + timedelta(days=random.randint(2, 5))
    next_billing_date = end_date + timedelta(days=random.randint(5, 10))

    return start_date, end_date, next_billing_date, transactions, current_balance, ground_truth


def fill_header(ws, statement_number, person_data):
    """Fill the header section of the bank statement."""
    ws["B2"] = "MasterCard"
    ws["B3"] = person_data["card_number"]
    ws["C2"] = person_data["first_name"]
    ws["C3"] = person_data["last_name"]
    ws["F2"] = statement_number
    ws["F3"] = 1

def fill_balances_and_dates(ws, start_date, end_date, next_billing_date, final_balance):
    """Fill balance and date information."""
    ws[f"D{START_BALANCE_ROW}"] = f"KONTOSTAND AM {start_date.strftime('%d.%m.%Y')}"
    ws[f"E{START_BALANCE_ROW}"].alignment = Alignment(horizontal='right')

    ws[f"D{END_BALANCE_ROW}"] = f"KONTOSTAND AM {end_date.strftime('%d.%m.%Y')}"
    ws[f"E{END_BALANCE_ROW}"] = f"{final_balance:.2f}".replace('.', ',') + "-"
    ws[f"E{END_BALANCE_ROW}"].alignment = Alignment(horizontal='right')

    ws[f"C{NEXT_BILLING_ROW}"] = f"IHR NAECHSTER ABRECHNUNGSTERMIN {next_billing_date.strftime('%d.%m.%Y')}"

def fill_transactions(ws, transactions):
    """Fill transaction data into the worksheet."""
    for i, trx_data in enumerate(transactions):
        row = TRANSACTION_START_ROW + i
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'] = trx_data
        ws[f'E{row}'].alignment = Alignment(horizontal='right')

    # Clear unused rows
    for i in range(len(transactions), MAX_TRANSACTIONS):
        row = TRANSACTION_START_ROW + i
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].value = None

def create_bank_statement(output_path, statement_number, person_data, fake):
    """
    Create a single bank statement Excel file from a template and return its ground truth data.
    """
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    start_date, end_date, next_billing_date, transactions, final_balance, ground_truth = generate_statement_data(fake, statement_number)

    fill_header(ws, statement_number, person_data)
    fill_balances_and_dates(ws, start_date, end_date, next_billing_date, final_balance)
    fill_transactions(ws, transactions)

    wb.save(output_path)
    print(f"✅ Successfully created Excel file: {output_path}")
    return ground_truth


def convert_excel_to_image(excel_path, image_path):
    """Convert Excel file to PNG image."""
    try:
        image_range = f"A1:F{NEXT_BILLING_ROW + 1}"
        wb = load_workbook(excel_path)
        ws_title = wb.active.title
        excel2img.export_img(str(excel_path), str(image_path), ws_title, image_range)
        print(f"📸 Successfully saved image: {image_path}")
        return True
    except Exception as e:
        print(f"❌ Error converting {excel_path} to image: {e}")
        return False

def save_ground_truth_data(all_ground_truth_data):
    """Save ground truth data to CSV file."""
    if not all_ground_truth_data:
        return

    df_ground_truth = pd.DataFrame(all_ground_truth_data)
    df_ground_truth.to_csv(GROUND_TRUTH_CSV, index=False, encoding='utf-8')
    print(f"💾 Ground truth data saved to: {GROUND_TRUTH_CSV}")

def main():
    """
    Main function to generate bank statements, convert them to images,
    and save a ground truth CSV file.
    """
    if not TEMPLATE_FILE.exists():
        print(f"❌ Error: Template file '{TEMPLATE_FILE}' not found.")
        return

    try:
        num_statements = int(input("How many bank statements do you want to generate? "))
        if num_statements <= 0:
            print("❌ Number must be positive.")
            return
    except ValueError:
        print("❌ Invalid number. Please enter a valid integer.")
        return

    print(f"🚀 Starting bank statement generation for {num_statements} statements...")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    fake = Faker('de_DE')
    person_data = {
        "first_name": fake.first_name(),
        "last_name": fake.last_name(),
        "card_number": fake.credit_card_number(card_type='mastercard'),
    }

    all_ground_truth_data = []

    for i in range(1, num_statements + 1):
        excel_path = OUTPUT_DIR / f"statement_{i}.xlsx"
        image_path = OUTPUT_DIR / f"statement_{i}.png"

        ground_truth = create_bank_statement(excel_path, i, person_data, fake)
        all_ground_truth_data.extend(ground_truth)

        convert_excel_to_image(excel_path, image_path)
        print()

    save_ground_truth_data(all_ground_truth_data)
    print(f"🎉 All {num_statements} statements generated in the '{OUTPUT_DIR}' folder.")


if __name__ == "__main__":
    main()
