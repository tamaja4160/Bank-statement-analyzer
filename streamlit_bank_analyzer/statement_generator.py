"""
Modified statement generator for Streamlit app
Generates bank statements with custom names and returns image paths.
"""

import os
import random
from datetime import datetime, timedelta
from pathlib import Path
import uuid

import excel2img
import pandas as pd
from faker import Faker
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Import comprehensive transaction data
try:
    # Try relative import first (for when imported as module)
    from .transaction_data import RECURRING_PAYMENTS, ONE_OFF_PAYMENTS, GERMAN_CITIES
except ImportError:
    # Fall back to absolute import (for when run directly)
    from transaction_data import RECURRING_PAYMENTS, ONE_OFF_PAYMENTS, GERMAN_CITIES

TRANSACTION_START_ROW = 6
TRANSACTION_END_ROW = 11
END_BALANCE_ROW = 12
NEXT_BILLING_ROW = 13
MAX_TRANSACTIONS_PER_STATEMENT = (TRANSACTION_END_ROW - TRANSACTION_START_ROW) + 1

def generate_statement_data(fake, statement_number, start_date, initial_balance, paid_subscriptions_this_month, current_month):
    """Generate realistic transaction data for a statement with continuity."""
    transactions = []
    current_balance = initial_balance
    current_date = start_date

    # Determine how many days this statement should cover (10-15 days)
    statement_days = random.randint(10, 15)
    end_date = start_date + timedelta(days=statement_days)

    # Get available recurring payments for this month
    available_recurring = []
    for payment in RECURRING_PAYMENTS:
        subscription_key = f"{payment['partner']}_{current_month}"
        if subscription_key not in paid_subscriptions_this_month:
            available_recurring.append(payment)

    # Calculate how many transactions we can fit
    max_transactions = MAX_TRANSACTIONS_PER_STATEMENT

    # Always include salary if it's the salary day and we haven't paid it this month
    salary_day = 25
    salary_amount = random.uniform(2500, 4000)  # Monthly salary range
    include_salary = (start_date.day <= salary_day <= end_date.day and
                     f"SALARY_{current_month}" not in paid_subscriptions_this_month)

    if include_salary:
        max_transactions -= 1  # Reserve space for salary

    # Ensure 2-3 mandatory subscriptions per month from essential categories
    essential_categories = ["Liability Insurance", "Car Insurance", "Home Insurance", "Electricity", "Gas", "Water", "Internet", "Mobile"]
    mandatory_payments = []

    # Try to get at least one from each essential category
    for category in essential_categories:
        category_payments = [p for p in available_recurring if p["category"] == category]
        if category_payments:
            # Select one payment from this category
            selected_payment = random.choice(category_payments)
            mandatory_payments.append(selected_payment)
            available_recurring.remove(selected_payment)

    # If we have too many mandatory payments, reduce to 2-3
    if len(mandatory_payments) > 3:
        mandatory_payments = random.sample(mandatory_payments, random.randint(2, 3))
    elif len(mandatory_payments) < 2 and available_recurring:
        # If we don't have enough mandatory, add some optional ones
        additional_needed = 2 - len(mandatory_payments)
        if available_recurring:
            additional_payments = random.sample(available_recurring[:additional_needed], min(additional_needed, len(available_recurring)))
            mandatory_payments.extend(additional_payments)
            for payment in additional_payments:
                available_recurring.remove(payment)

    # Select additional recurring payments if space allows
    remaining_slots = max_transactions - len(mandatory_payments)
    if remaining_slots > 0 and available_recurring:
        num_additional = min(random.randint(0, remaining_slots), len(available_recurring))
        additional_recurring = random.sample(available_recurring, num_additional)
        mandatory_payments.extend(additional_recurring)

    selected_recurring = mandatory_payments

    # Calculate remaining slots for one-off payments
    remaining_slots = max_transactions - len(selected_recurring) - (1 if include_salary else 0)
    num_one_off = random.randint(0, min(remaining_slots, len(ONE_OFF_PAYMENTS)))

    # Create transaction pool
    transaction_pool = selected_recurring + [None] * num_one_off
    random.shuffle(transaction_pool)

    # Generate transactions
    for transaction_type in transaction_pool:
        # Space out transactions across the statement period
        days_to_add = random.randint(1, max(1, statement_days // len(transaction_pool)))
        current_date = min(current_date + timedelta(days=days_to_add), end_date)

        beleg_date = current_date.strftime("%d.%m.")
        valuta_date = (current_date + timedelta(days=1)).strftime("%d.%m.")

        if transaction_type:  # It's a recurring payment
            partner = transaction_type["partner"]
            description = transaction_type["description_template"].format(
                partner=partner,
                contract_id=fake.random_number(digits=8)
            )
            amount = round(transaction_type["base_amount"] + random.uniform(
                transaction_type["variation"][0], transaction_type["variation"][1]), 2)

            # Mark this subscription as paid for this month
            subscription_key = f"{partner}_{current_month}"
            paid_subscriptions_this_month.add(subscription_key)
        else:  # It's a one-off payment
            one_off = random.choice(ONE_OFF_PAYMENTS)
            partner = random.choice(one_off["partner_options"])
            description = one_off["description_template"].format(
                date_short=current_date.strftime("%d.%m"),
                partner=partner,
                random_str=fake.lexify(text='??????').upper(),
                city=random.choice(GERMAN_CITIES)
            )
            amount = round(random.uniform(*one_off["amount_range"]), 2)

        transactions.append([beleg_date, valuta_date, description, f"{amount:.2f}".replace('.', ',') + "-"])
        current_balance += amount

    # Add salary payment if applicable
    if include_salary:
        salary_date = start_date.replace(day=salary_day)
        if start_date <= salary_date <= end_date:
            beleg_date = salary_date.strftime("%d.%m.")
            valuta_date = (salary_date + timedelta(days=1)).strftime("%d.%m.")
            description = "GEHALT MONATLICH"
            amount = round(salary_amount, 2)

            transactions.append([beleg_date, valuta_date, description, f"{amount:.2f}".replace('.', ',')])
            current_balance += amount

            # Mark salary as paid for this month
            paid_subscriptions_this_month.add(f"SALARY_{current_month}")

    # Sort transactions by date
    transactions.sort(key=lambda x: datetime.strptime(x[0], "%d.%m."))

    next_billing_date = end_date + timedelta(days=random.randint(5, 10))

    return end_date, next_billing_date, transactions, current_balance, paid_subscriptions_this_month

def generate_statements_for_name(name: str, num_statements: int = 10) -> list:
    """
    Generate bank statements for a specific name with continuity and return image paths.

    Args:
        name: Full name to use for the statements
        num_statements: Number of statements to generate

    Returns:
        List of paths to generated PNG images
    """
    try:
        # Create session-specific output directory
        session_id = str(uuid.uuid4())[:8]
        output_dir = Path(f"streamlit_sessions/{session_id}")
        output_dir.mkdir(parents=True, exist_ok=True)

        # Split name into first and last
        name_parts = name.strip().split()
        if len(name_parts) >= 2:
            first_name = name_parts[0]
            last_name = ' '.join(name_parts[1:])
        else:
            first_name = name_parts[0]
            last_name = "Doe"  # Default last name

        # Generate fake data but use provided name
        fake = Faker('de_DE')

        person_data = {
            "first_name": first_name,
            "last_name": last_name,
            "card_number": fake.credit_card_number(card_type='mastercard'),
        }

        # Initialize continuity variables
        initial_balance = round(random.uniform(1000, 4500), 2)  # Random starting balance (1000-4500€)
        current_date = fake.date_between(start_date="-1y", end_date="-3M")  # Start 3-12 months ago
        paid_subscriptions_this_month = set()
        current_month = current_date.strftime("%Y-%m")
        running_balance = initial_balance

        image_paths = []

        for i in range(1, num_statements + 1):
            excel_filename = output_dir / f"statement_{i}.xlsx"
            image_filename = output_dir / f"statement_{i}.png"

            # Check if we've moved to a new month
            statement_month = current_date.strftime("%Y-%m")
            if statement_month != current_month:
                current_month = statement_month
                paid_subscriptions_this_month.clear()  # Reset subscriptions for new month

            # Generate statement data with continuity
            end_date, next_billing_date, transactions, final_balance, paid_subscriptions_this_month = generate_statement_data(
                fake, i, current_date, running_balance, paid_subscriptions_this_month, current_month
            )

            # Create Excel file
            template_path = os.path.join(os.path.dirname(__file__), "..", "bank statement generator", "template.xlsx")
            wb = load_workbook(template_path)
            ws = wb.active

            # Fill header with consistent person data
            ws["B2"] = "MasterCard"
            ws["B3"] = person_data["card_number"]
            ws["C2"] = person_data["first_name"]
            ws["C3"] = person_data["last_name"]
            ws["F2"] = i
            ws["F3"] = 1

            # Fill fixed position balances and dates
            ws[f"D{5}"] = f"KONTOSTAND AM {current_date.strftime('%d.%m.%Y')}"
            ws[f"E{5}"] = f"{abs(running_balance):.2f}".replace('.', ',')
            ws[f"E{5}"].alignment = Alignment(horizontal='right')

            ws[f"D{12}"] = f"KONTOSTAND AM {end_date.strftime('%d.%m.%Y')}"
            ws[f"E{12}"] = f"{abs(final_balance):.2f}".replace('.', ',')
            ws[f"E{12}"].alignment = Alignment(horizontal='right')

            ws[f"C{13}"] = f"IHR NAECHSTER ABRECHNUNGSTERMIN {next_billing_date.strftime('%d.%m.%Y')}"

            # Fill transaction rows
            for j, trx_data in enumerate(transactions):
                row = TRANSACTION_START_ROW + j
                ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'] = trx_data
                ws[f'E{row}'].alignment = Alignment(horizontal='right')

            # Clear any unused transaction rows
            for j in range(len(transactions), MAX_TRANSACTIONS_PER_STATEMENT):
                row_to_clear = TRANSACTION_START_ROW + j
                ws[f'B{row_to_clear}'].value, ws[f'C{row_to_clear}'].value, ws[f'D{row_to_clear}'].value, ws[f'E{row_to_clear}'].value = (None, None, None, None)

            wb.save(excel_filename)

            # Convert to image
            try:
                image_range = f"A1:F{NEXT_BILLING_ROW + 1}"
                excel2img.export_img(str(excel_filename), str(image_filename), wb.active.title, image_range)
                image_paths.append(str(image_filename))
            except Exception as e:
                print(f"Error converting {excel_filename} to image: {e}")
                continue

            # Update for next statement
            current_date = end_date + timedelta(days=1)  # Next statement starts day after this one ends
            running_balance = final_balance

        return image_paths

    except Exception as e:
        print(f"Error generating statements: {e}")
        return []
