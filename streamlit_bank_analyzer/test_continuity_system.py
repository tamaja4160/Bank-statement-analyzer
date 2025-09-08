"""
Test suite for the bank statement continuity system.
Tests realistic scenario requirements including sequential generation,
monthly subscription tracking, balance continuity, and salary integration.
"""

import os
import sys
import tempfile
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import openpyxl

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(__file__))

from statement_generator import generate_statements_for_name
from transaction_data import RECURRING_PAYMENTS

class ContinuitySystemTester:
    """Test class for validating the continuity system requirements."""

    def __init__(self):
        self.test_results = []
        self.temp_dir = None

    def setup_test_environment(self):
        """Set up temporary directory for testing."""
        self.temp_dir = tempfile.mkdtemp()
        # Change to temp directory to avoid conflicts with existing sessions
        os.chdir(self.temp_dir)

    def cleanup_test_environment(self):
        """Clean up temporary test files."""
        if self.temp_dir and os.path.exists(self.temp_dir):
            os.chdir(os.path.dirname(__file__))  # Go back to original directory

            # Try to clean up with retries for locked files
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    shutil.rmtree(self.temp_dir)
                    break
                except PermissionError:
                    if attempt < max_retries - 1:
                        print(f"  ⚠️  Cleanup attempt {attempt + 1} failed, retrying...")
                        import time
                        time.sleep(1)  # Wait 1 second before retry
                    else:
                        print(f"  ⚠️  Could not clean up temporary directory: {self.temp_dir}")
                        print("    This doesn't affect test results, files will be cleaned up later.")

    def run_all_tests(self):
        """Run all continuity system tests."""
        print("🧪 Running Continuity System Tests...")
        print("=" * 50)

        try:
            self.setup_test_environment()

            # Test 1: Sequential Statement Generation
            self.test_sequential_statements()

            # Test 2: Monthly Subscription Tracking
            self.test_monthly_subscription_tracking()

            # Test 3: Account Balance Continuity
            self.test_balance_continuity()

            # Test 4: Monthly Salary Integration
            self.test_monthly_salary_integration()

            # Test 5: Time Period Management
            self.test_time_period_management()

            # Test 6: No Duplicate Subscriptions
            self.test_no_duplicate_subscriptions()

            self.print_test_summary()

        except Exception as e:
            print(f"❌ Test execution failed: {e}")
            import traceback
            traceback.print_exc()
        finally:
            self.cleanup_test_environment()

    def test_sequential_statements(self):
        """Test that statements are generated sequentially."""
        print("\n📅 Test 1: Sequential Statement Generation")

        # Generate test statements
        image_paths = generate_statements_for_name("Test User", 3)

        if len(image_paths) != 3:
            self.log_test_result("Sequential Statements", False, f"Expected 3 statements, got {len(image_paths)}")
            return

        # Extract session directory from first path
        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        if len(excel_files) != 3:
            self.log_test_result("Sequential Statements", False, f"Expected 3 Excel files, got {len(excel_files)}")
            return

        # Check date continuity
        previous_end_date = None
        for i, excel_file in enumerate(excel_files):
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            # Get start and end dates
            start_balance_text = ws['D5'].value
            end_balance_text = ws['D12'].value

            if "KONTOSTAND AM" in str(start_balance_text):
                start_date_str = start_balance_text.replace("KONTOSTAND AM ", "")
                start_date = datetime.strptime(start_date_str, "%d.%m.%Y")

            if "KONTOSTAND AM" in str(end_balance_text):
                end_date_str = end_balance_text.replace("KONTOSTAND AM ", "")
                end_date = datetime.strptime(end_date_str, "%d.%m.%Y")

            # Check that this statement starts the day after the previous one ended
            if previous_end_date and start_date != previous_end_date + timedelta(days=1):
                self.log_test_result("Sequential Statements", False,
                                   f"Statement {i+1}: Expected start date {previous_end_date + timedelta(days=1)}, got {start_date}")
                return

            previous_end_date = end_date

        self.log_test_result("Sequential Statements", True, "All statements generated sequentially")

    def test_monthly_subscription_tracking(self):
        """Test that subscriptions are tracked per month."""
        print("\n📊 Test 2: Monthly Subscription Tracking")

        # Generate statements that span multiple months
        image_paths = generate_statements_for_name("Test User", 8)

        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        monthly_subscriptions = {}
        salary_payments = {}

        for excel_file in excel_files:
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            # Extract transactions
            transactions = []
            for row in range(6, 12):  # Transaction rows
                beleg_date = ws[f'B{row}'].value
                description = ws[f'D{row}'].value
                amount = ws[f'E{row}'].value

                if beleg_date and description:
                    transactions.append({
                        'date': beleg_date,
                        'description': description,
                        'amount': amount
                    })

            # Analyze transactions for the month
            for transaction in transactions:
                if transaction['date']:
                    try:
                        date_obj = datetime.strptime(transaction['date'], "%d.%m.")
                        month_key = date_obj.strftime("%Y-%m")

                        # Track salary payments
                        if "GEHALT" in transaction['description']:
                            if month_key not in salary_payments:
                                salary_payments[month_key] = []
                            salary_payments[month_key].append(transaction)

                        # Track subscriptions
                        for payment in RECURRING_PAYMENTS:
                            if payment['partner'] in transaction['description']:
                                if month_key not in monthly_subscriptions:
                                    monthly_subscriptions[month_key] = {}
                                if payment['partner'] not in monthly_subscriptions[month_key]:
                                    monthly_subscriptions[month_key][payment['partner']] = []
                                monthly_subscriptions[month_key][payment['partner']].append(transaction)
                                break
                    except ValueError:
                        continue

        # Check that no subscription appears more than once per month
        duplicates_found = False
        for month, subscriptions in monthly_subscriptions.items():
            for partner, transactions in subscriptions.items():
                if len(transactions) > 1:
                    duplicates_found = True
                    print(f"  ⚠️  {partner} appears {len(transactions)} times in {month}")

        # Check salary payments (should be 0 or 1 per month)
        salary_issues = False
        for month, payments in salary_payments.items():
            if len(payments) > 1:
                salary_issues = True
                print(f"  ⚠️  Salary paid {len(payments)} times in {month}")

        if duplicates_found or salary_issues:
            self.log_test_result("Monthly Subscription Tracking", False,
                               "Found duplicate subscriptions or multiple salary payments in same month")
        else:
            self.log_test_result("Monthly Subscription Tracking", True,
                               "No duplicate subscriptions or salary payments found")

    def test_balance_continuity(self):
        """Test that account balance flows correctly between statements."""
        print("\n💰 Test 3: Account Balance Continuity")

        image_paths = generate_statements_for_name("Test User", 4)

        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        previous_end_balance = None

        for i, excel_file in enumerate(excel_files):
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            # Get start and end balances
            start_balance_text = ws['E5'].value
            end_balance_text = ws['E12'].value

            if start_balance_text:
                start_balance = float(start_balance_text.replace('-', '').replace(',', '.'))
            if end_balance_text:
                end_balance = float(end_balance_text.replace('-', '').replace(',', '.'))

            # Check balance continuity
            if previous_end_balance is not None and abs(start_balance - previous_end_balance) > 0.01:
                self.log_test_result("Balance Continuity", False,
                                   f"Statement {i+1}: Balance discontinuity - expected {previous_end_balance}, got {start_balance}")
                return

            previous_end_balance = end_balance

        self.log_test_result("Balance Continuity", True, "Account balance flows correctly between statements")

    def test_monthly_salary_integration(self):
        """Test that salary is paid monthly on the 25th."""
        print("\n💼 Test 4: Monthly Salary Integration")

        # Generate enough statements to cover multiple months
        image_paths = generate_statements_for_name("Test User", 10)

        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        salary_dates = []
        salary_amounts = []

        for excel_file in excel_files:
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            # Extract transactions
            for row in range(6, 12):
                beleg_date = ws[f'B{row}'].value
                description = ws[f'D{row}'].value
                amount = ws[f'E{row}'].value

                if beleg_date and description and "GEHALT" in description:
                    try:
                        date_obj = datetime.strptime(beleg_date, "%d.%m.")
                        salary_dates.append(date_obj.day)
                        if amount:
                            salary_amount = float(amount.replace(',', '.'))
                            salary_amounts.append(salary_amount)
                    except ValueError:
                        continue

        # Check that salary is paid on the 25th
        if salary_dates and not all(date == 25 for date in salary_dates):
            self.log_test_result("Monthly Salary Integration", False,
                               f"Salary not always paid on 25th. Dates found: {salary_dates}")
            return

        # Check salary amounts are reasonable (€2,500-€4,000)
        if salary_amounts and not all(2500 <= abs(amount) <= 4000 for amount in salary_amounts):
            self.log_test_result("Monthly Salary Integration", False,
                               f"Salary amounts out of range: {salary_amounts}")
            return

        self.log_test_result("Monthly Salary Integration", True,
                           f"Salary correctly paid on 25th with amounts: {salary_amounts}")

    def test_time_period_management(self):
        """Test that statements have proper time periods."""
        print("\n⏰ Test 5: Time Period Management")

        image_paths = generate_statements_for_name("Test User", 5)

        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        previous_end_date = None

        for i, excel_file in enumerate(excel_files):
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            # Get start and end dates
            start_balance_text = ws['D5'].value
            end_balance_text = ws['D12'].value

            if "KONTOSTAND AM" in str(start_balance_text):
                start_date_str = start_balance_text.replace("KONTOSTAND AM ", "")
                start_date = datetime.strptime(start_date_str, "%d.%m.%Y")

            if "KONTOSTAND AM" in str(end_balance_text):
                end_date_str = end_balance_text.replace("KONTOSTAND AM ", "")
                end_date = datetime.strptime(end_date_str, "%d.%m.%Y")

            # Check statement duration (should be 10-15 days)
            duration = (end_date - start_date).days
            if not (10 <= duration <= 15):
                self.log_test_result("Time Period Management", False,
                                   f"Statement {i+1}: Duration {duration} days, expected 10-15 days")
                return

            # Check chronological order
            if previous_end_date and start_date <= previous_end_date:
                self.log_test_result("Time Period Management", False,
                                   f"Statement {i+1}: Start date {start_date} not after previous end date {previous_end_date}")
                return

            previous_end_date = end_date

        self.log_test_result("Time Period Management", True, "All statements have proper time periods")

    def test_no_duplicate_subscriptions(self):
        """Test that no subscription appears more than once per month."""
        print("\n🚫 Test 6: No Duplicate Subscriptions")

        # This is already tested in test_monthly_subscription_tracking
        # but we'll run it again with a focused test
        image_paths = generate_statements_for_name("Test User", 6)

        session_dir = os.path.dirname(image_paths[0])
        excel_files = [f for f in os.listdir(session_dir) if f.endswith('.xlsx')]
        excel_files.sort()

        subscription_counts = {}

        for excel_file in excel_files:
            wb = openpyxl.load_workbook(os.path.join(session_dir, excel_file))
            ws = wb.active

            for row in range(6, 12):
                beleg_date = ws[f'B{row}'].value
                description = ws[f'D{row}'].value

                if beleg_date and description:
                    try:
                        date_obj = datetime.strptime(beleg_date, "%d.%m.")
                        month_key = date_obj.strftime("%Y-%m")

                        for payment in RECURRING_PAYMENTS:
                            if payment['partner'] in description:
                                key = f"{month_key}_{payment['partner']}"
                                if key not in subscription_counts:
                                    subscription_counts[key] = 0
                                subscription_counts[key] += 1
                                break
                    except ValueError:
                        continue

        # Check for duplicates
        duplicates = [key for key, count in subscription_counts.items() if count > 1]

        if duplicates:
            self.log_test_result("No Duplicate Subscriptions", False,
                               f"Found duplicate subscriptions: {duplicates}")
        else:
            self.log_test_result("No Duplicate Subscriptions", True,
                               "No duplicate subscriptions found within months")

    def log_test_result(self, test_name, passed, message):
        """Log a test result."""
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"  {status}: {message}")
        self.test_results.append({
            'test': test_name,
            'passed': passed,
            'message': message
        })

    def print_test_summary(self):
        """Print a summary of all test results."""
        print("\n" + "=" * 50)
        print("📊 TEST SUMMARY")
        print("=" * 50)

        passed_tests = sum(1 for result in self.test_results if result['passed'])
        total_tests = len(self.test_results)

        for result in self.test_results:
            status = "✅" if result['passed'] else "❌"
            print(f"{status} {result['test']}: {result['message']}")

        print(f"\n🎯 Overall: {passed_tests}/{total_tests} tests passed")

        if passed_tests == total_tests:
            print("🎉 All tests passed! Continuity system is working correctly.")
        else:
            print("⚠️  Some tests failed. Please review the continuity system implementation.")

def main():
    """Run the continuity system tests."""
    tester = ContinuitySystemTester()
    tester.run_all_tests()

if __name__ == "__main__":
    main()
