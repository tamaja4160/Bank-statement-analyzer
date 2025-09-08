import os
import random
from datetime import datetime, timedelta

import excel2img
import pandas as pd
from faker import Faker
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- CONFIGURATION ---
NUM_STATEMENTS_TO_GENERATE = 20  # Reduced for faster generation
OUTPUT_FOLDER = "bank statement generator/bank_statements"
TEMPLATE_FILE = "bank statement generator/template.xlsx"
GROUND_TRUTH_CSV = os.path.join(OUTPUT_FOLDER, "ground_truth.csv")

# --- TEMPLATE LAYOUT CONSTANTS ---
START_BALANCE_ROW = 5
TRANSACTION_START_ROW = 6
TRANSACTION_END_ROW = 11
END_BALANCE_ROW = 12
NEXT_BILLING_ROW = 13
MAX_TRANSACTIONS_PER_STATEMENT = (TRANSACTION_END_ROW - TRANSACTION_START_ROW) + 1

# --- DATA FOR RECURRING PAYMENTS ---
ZEUS_BODYPOWER_PARTNER = "ZEUS BODYPOWER"
ZEUS_BODYPOWER_VARIATION_AMOUNT = 0.5  # Maximum variation for ZEUS BODYPOWER payments
ZEUS_BODYPOWER_VARIATION = (-ZEUS_BODYPOWER_VARIATION_AMOUNT, ZEUS_BODYPOWER_VARIATION_AMOUNT)

RECURRING_PAYMENTS = [
    {
        "partner": "Allianz SE", "category": "Insurance",
        "description_template": "BEITRAG {partner} K-{contract_id}",
        "base_amount": 55.20
    },
    {
        "partner": "VODAFONE GMBH", "category": "Internet",
        "description_template": "RECHNUNG {partner} {contract_id}",
        "base_amount": 39.99
    },
    {
        "partner": ZEUS_BODYPOWER_PARTNER, "category": "Gym",
        "description_template": "MITGLIEDSBEITRAG {partner}",
        "base_amount": 25.00
    },
    {
        "partner": "Stadtwerke Rosenheim", "category": "Electricity",
        "description_template": "ABSCHLAG STROM {partner} {contract_id}",
        "base_amount": 85.50
    },
]

# --- NEW: Pool of realistic random, one-off payments ---
ONE_OFF_PAYMENTS = [
    {
        "partner_options": ["EDEKA", "REWE", "LIDL", "ALDI SUED"], "category": "Groceries",
        "description_template": "KARTENZ./{date_short} {partner} RO",
        "amount_range": (15.0, 150.0)
    },
    {
        "partner_options": ["AMAZON.DE", "ZALANDO", "EBAY"], "category": "Shopping",
        "description_template": "{partner} MKTPLC EU {random_str}",
        "amount_range": (10.0, 250.0)
    },
    {
        "partner_options": ["SHELL", "ARAL", "JET"], "category": "Fuel",
        "description_template": "KARTENZAHLUNG {partner} TANKSTELLE",
        "amount_range": (40.0, 90.0)
    },
    {
        "partner_options": ["BURGER KING", "MCDONALDS"], "category": "Dining",
        "description_template": "{partner} {city}",
        "amount_range": (8.0, 45.0)
    },
    {
       "partner_options": ["PAYPAL"], "category": "General",
       "description_template": "PAYPAL {random_str}",
       "amount_range": (5.0, 100.0)
    }
]


def generate_statement_data(fake: Faker, statement_number: int):
    """
    Generates realistic transaction data, creating a mix of recurring and
    varied one-off payments.
    """
    transactions = []
    ground_truth_for_statement = []
    current_balance = 0.0
    start_date = fake.date_between(start_date="-2y", end_date="today")
    current_date = start_date

    # --- CHANGED: Generate a more varied mix of transactions ---
    num_total_transactions = random.randint(3, MAX_TRANSACTIONS_PER_STATEMENT)

    # Ensure Vodafone appears in at least 60% of statements for testing
    guaranteed_recurring = []
    if random.random() < 0.6:  # 60% chance
        vodafone_payment = next(p for p in RECURRING_PAYMENTS if p["partner"] == "VODAFONE GMBH")
        guaranteed_recurring.append(vodafone_payment)

    # Add other recurring payments randomly
    remaining_payments = [p for p in RECURRING_PAYMENTS if p["partner"] != "VODAFONE GMBH"]
    if remaining_payments:
        num_additional = random.randint(0, min(len(remaining_payments), num_total_transactions - len(guaranteed_recurring) - 1))
        if num_additional > 0:
            additional_recurring = random.sample(remaining_payments, num_additional)
            guaranteed_recurring.extend(additional_recurring)

    potential_recurring = guaranteed_recurring
    
    transaction_pool = potential_recurring + [None] * (num_total_transactions - len(potential_recurring))
    random.shuffle(transaction_pool)

    for transaction_type in transaction_pool:
        current_date += timedelta(days=random.randint(1, 4))
        beleg_date = current_date.strftime("%d.%m.")
        valuta_date = (current_date + timedelta(days=1)).strftime("%d.%m.")

        if transaction_type:  # It's a recurring payment
            partner = transaction_type["partner"]
            category = transaction_type["category"]
            description = transaction_type["description_template"].format(
                partner=partner,
                contract_id=fake.random_number(digits=8)
            )

            # Special handling for ZEUS BODYPOWER - much less variation
            if partner == ZEUS_BODYPOWER_PARTNER:
                # Only ±0.5€ variation for gym membership (much more consistent)
                amount = round(transaction_type["base_amount"] + random.uniform(*ZEUS_BODYPOWER_VARIATION), 2)
            else:
                # Original variation for other recurring payments
                amount = round(transaction_type["base_amount"] + random.uniform(-2.5, 2.5), 2)
        else:  # It's a one-off payment from the new pool
            one_off = random.choice(ONE_OFF_PAYMENTS)
            partner = random.choice(one_off["partner_options"])
            category = one_off["category"]
            description = one_off["description_template"].format(
                date_short=current_date.strftime("%d.%m"),
                partner=partner,
                random_str=fake.lexify(text='??????').upper(),
                city=fake.city()
            )
            amount = round(random.uniform(*one_off["amount_range"]), 2)

        transactions.append([beleg_date, valuta_date, description, f"{amount:.2f}".replace('.', ',') + "-"])
        current_balance += amount

        ground_truth_for_statement.append({
            "statement_id": statement_number,
            "date": current_date.strftime("%d.%m.%Y"),
            "description": description,
            "amount": -amount,
            "identified_partner": partner,
            "category": category
        })

    end_date = current_date + timedelta(days=random.randint(2, 5))
    next_billing_date = end_date + timedelta(days=random.randint(5, 10))

    return start_date, end_date, next_billing_date, transactions, current_balance, ground_truth_for_statement


def create_bank_statement(output_path: str, statement_number: int, person_data: dict, fake: Faker):
    """
    Creates a single bank statement Excel file from a template and returns its ground truth data.
    """
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    start_date, end_date, next_billing_date, transactions, final_balance, ground_truth = generate_statement_data(fake, statement_number)

    # Fill header with consistent person data
    ws["B2"] = "MasterCard"
    ws["B3"] = person_data["card_number"]
    ws["C2"] = person_data["first_name"]
    ws["C3"] = person_data["last_name"]
    ws["F2"] = statement_number
    ws["F3"] = 1

    # Fill fixed position balances and dates
    ws[f"D{START_BALANCE_ROW}"] = f"KONTOSTAND AM {start_date.strftime('%d.%m.%Y')}"
    ws[f"E{START_BALANCE_ROW}"].alignment = Alignment(horizontal='right')
    
    ws[f"D{END_BALANCE_ROW}"] = f"KONTOSTAND AM {end_date.strftime('%d.%m.%Y')}"
    ws[f"E{END_BALANCE_ROW}"] = f"{final_balance:.2f}".replace('.', ',') + "-"
    ws[f"E{END_BALANCE_ROW}"].alignment = Alignment(horizontal='right')
    
    ws[f"C{NEXT_BILLING_ROW}"] = f"IHR NAECHSTER ABRECHNUNGSTERMIN {next_billing_date.strftime('%d.%m.%Y')}"

    # Fill transaction rows
    for i, trx_data in enumerate(transactions):
        row = TRANSACTION_START_ROW + i
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'] = trx_data
        ws[f'E{row}'].alignment = Alignment(horizontal='right')

    # Clear any unused transaction rows
    for i in range(len(transactions), MAX_TRANSACTIONS_PER_STATEMENT):
        row_to_clear = TRANSACTION_START_ROW + i
        ws[f'B{row_to_clear}'].value, ws[f'C{row_to_clear}'].value, ws[f'D{row_to_clear}'].value, ws[f'E{row_to_clear}'].value = (None, None, None, None)

    wb.save(output_path)
    print(f"✅ Successfully created Excel file: {output_path}")
    return ground_truth


def main():
    """
    Main function to generate bank statements, convert them to images,
    and save a ground truth CSV file.
    """
    if not os.path.exists(TEMPLATE_FILE):
        print(f"❌ Error: Template file '{TEMPLATE_FILE}' not found.")
        return

    print("🚀 Starting bank statement generation...")
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    fake = Faker('de_DE')
    all_ground_truth_data = []

    person_data = {
        "first_name": fake.first_name(),
        "last_name": fake.last_name(),
        "card_number": fake.credit_card_number(card_type='mastercard'),
    }

    for i in range(1, NUM_STATEMENTS_TO_GENERATE + 1):
        excel_filename = os.path.join(OUTPUT_FOLDER, f"statement_{i}.xlsx")
        image_filename = os.path.join(OUTPUT_FOLDER, f"statement_{i}.png")
        
        statement_ground_truth = create_bank_statement(excel_filename, i, person_data, fake)
        all_ground_truth_data.extend(statement_ground_truth)
        
        try:
            image_range = f"A1:F{NEXT_BILLING_ROW + 1}"
            wb = load_workbook(excel_filename)
            ws_title = wb.active.title
            excel2img.export_img(excel_filename, image_filename, ws_title, image_range)
            print(f"📸 Successfully saved image: {image_filename}\n")
        except Exception as e:
            print(f"❌ Error converting {excel_filename} to image: {e}\n")
            
    if all_ground_truth_data:
        df_ground_truth = pd.DataFrame(all_ground_truth_data)
        df_ground_truth.to_csv(GROUND_TRUTH_CSV, index=False, encoding='utf-8')
        print(f"💾 Ground truth data saved to: {GROUND_TRUTH_CSV}")

    print(f"🎉 All {NUM_STATEMENTS_TO_GENERATE} statements generated in the '{OUTPUT_FOLDER}' folder.")


if __name__ == "__main__":
    main()
